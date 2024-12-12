# import streamlit as st
# import pandas as pd
# import openpyxl
# from datetime import datetime
# import os
#
#
# class TeamStatusTracker:
#     def __init__(self, file_path='team_status_tracker.xlsx'):
#         self.file_path = file_path
#         self.ensure_file_exists()
#
#     def ensure_file_exists(self):
#         """Create the Excel file if it doesn't exist"""
#         if not os.path.exists(self.file_path):
#             # Create a workbook with a default sheet
#             wb = openpyxl.Workbook()
#             wb.save(self.file_path)
#
#     def create_team_member_sheet(self, member_name):
#         """Create a new sheet for a team member"""
#         try:
#             # Read the existing workbook
#             book = openpyxl.load_workbook(self.file_path)
#
#             # Check if sheet already exists
#             if member_name in book.sheetnames:
#                 st.warning(f"Sheet for {member_name} already exists!")
#                 return False
#
#             # Create new sheet
#             book.create_sheet(title=member_name)
#
#             # Get the new sheet and add headers
#             sheet = book[member_name]
#             headers = ['Date', 'Tasks Completed', 'Tasks Pending', 'Blockers', 'Hours Worked']
#             for col, header in enumerate(headers, start=1):
#                 sheet.cell(row=1, column=col, value=header)
#
#             # Save the workbook
#             book.save(self.file_path)
#             st.success(f"Sheet created for {member_name}")
#             return True
#         except Exception as e:
#             st.error(f"Error creating sheet: {e}")
#             return False
#
#     def log_daily_status(self, member_name, tasks_completed, tasks_pending, blockers, hours_worked):
#         """Log daily status for a team member"""
#         try:
#             # Read the existing workbook
#             book = openpyxl.load_workbook(self.file_path)
#
#             # Check if sheet exists
#             if member_name not in book.sheetnames:
#                 st.error(f"Sheet for {member_name} does not exist. Create it first!")
#                 return False
#
#             # Get the sheet
#             sheet = book[member_name]
#
#             # Find the next row
#             next_row = sheet.max_row + 1
#
#             # Write data
#             sheet.cell(row=next_row, column=1, value=datetime.now().strftime('%Y-%m-%d'))
#             sheet.cell(row=next_row, column=2, value=tasks_completed)
#             sheet.cell(row=next_row, column=3, value=tasks_pending)
#             sheet.cell(row=next_row, column=4, value=blockers)
#             sheet.cell(row=next_row, column=5, value=hours_worked)
#
#             # Save the workbook
#             book.save(self.file_path)
#             st.success("Daily status logged successfully!")
#             return True
#         except Exception as e:
#             st.error(f"Error logging status: {e}")
#             return False
#
#
# def main():
#     # Page configuration
#     st.set_page_config(page_title="Team Status Tracker", page_icon="📊")
#
#     # Initialize the tracker
#     tracker = TeamStatusTracker()
#
#     # Title
#     st.title("🚀 Team Daily Status Tracker")
#
#     # Tabs for different functionalities
#     tab1, tab2 = st.tabs(["Create Team Member Sheet", "Log Daily Status"])
#
#     with tab1:
#         st.header("Create New Team Member Sheet")
#         new_member_name = st.text_input("Enter Team Member Name")
#         if st.button("Create Sheet"):
#             if new_member_name:
#                 tracker.create_team_member_sheet(new_member_name)
#             else:
#                 st.warning("Please enter a team member name")
#
#     with tab2:
#         st.header("Log Daily Status")
#
#         # Load existing sheets
#         try:
#             book = openpyxl.load_workbook(tracker.file_path)
#             existing_sheets = [sheet for sheet in book.sheetnames if sheet != 'Sheet']
#         except:
#             existing_sheets = []
#
#         # Member selection
#         member_name = st.selectbox("Select Team Member", existing_sheets)
#
#         # Status input fields
#         tasks_completed = st.text_area("Tasks Completed Today")
#         tasks_pending = st.text_area("Tasks Pending")
#         blockers = st.text_area("Any Blockers?")
#         hours_worked = st.number_input("Hours Worked Today", min_value=0.0, max_value=24.0, step=0.5)
#
#         # Log status button
#         if st.button("Log Status"):
#             if member_name:
#                 tracker.log_daily_status(
#                     member_name,
#                     tasks_completed,
#                     tasks_pending,
#                     blockers,
#                     hours_worked
#                 )
#             else:
#                 st.warning("Please select a team member first")
#
#     # Display current tracker location
#     st.sidebar.info(f"Tracker File: {os.path.abspath(tracker.file_path)}")
#
# #
# # if __name__ == "__main__":
# #     main()
#
#
# import streamlit as st
# import pandas as pd
# import openpyxl
# from datetime import datetime
# import os
# import hashlib
#
#
# class TeamStatusTracker:
#     def __init__(self, file_path='team_status_tracker.xlsx'):
#         self.file_path = file_path
#         self.ensure_file_exists()
#         self.users = self.load_users()
#
#     def ensure_file_exists(self):
#         """Create the Excel file if it doesn't exist"""
#         if not os.path.exists(self.file_path):
#             wb = openpyxl.Workbook()
#             wb.save(self.file_path)
#
#         # Ensure users sheet exists
#         try:
#             book = openpyxl.load_workbook(self.file_path)
#             if 'Users' not in book.sheetnames:
#                 users_sheet = book.create_sheet('Users')
#                 users_sheet['A1'] = 'Username'
#                 users_sheet['B1'] = 'Password'
#                 book.save(self.file_path)
#         except Exception as e:
#             st.error(f"Error creating users sheet: {e}")
#
#     def hash_password(self, password):
#         """Hash password for secure storage"""
#         return hashlib.sha256(password.encode()).hexdigest()
#
#     def load_users(self):
#         """Load users from the Excel file"""
#         try:
#             book = openpyxl.load_workbook(self.file_path)
#             users_sheet = book['Users']
#
#             users = {}
#             for row in users_sheet.iter_rows(min_row=2, values_only=True):
#                 if row[0] and row[1]:
#                     users[row[0]] = row[1]
#             return users
#         except Exception as e:
#             st.error(f"Error loading users: {e}")
#             return {}
#
#     def register_user(self, username, password):
#         """Register a new user"""
#         try:
#             # Check for empty username or password
#             if not username or not password:
#                 st.warning("Username and password cannot be empty!")
#                 return False
#
#             # Check if username already exists
#             if username in self.users:
#                 st.warning("Username already exists!")
#                 return False
#
#             # Open workbook
#             book = openpyxl.load_workbook(self.file_path)
#             users_sheet = book['Users']
#
#             # Find next empty row
#             next_row = users_sheet.max_row + 1
#
#             # Add new user
#             users_sheet.cell(row=next_row, column=1, value=username)
#             users_sheet.cell(row=next_row, column=2, value=self.hash_password(password))
#
#             # Save and update
#             book.save(self.file_path)
#             self.users[username] = self.hash_password(password)
#
#             st.success("User registered successfully!")
#             return True
#         except Exception as e:
#             st.error(f"Error registering user: {e}")
#             return False
#
#     def validate_login(self, username, password):
#         """Validate user login"""
#         if username in self.users:
#             return self.users[username] == self.hash_password(password)
#         return False
#
#
# def main():
#     # Initialize session state
#     if 'tracker' not in st.session_state:
#         st.session_state.tracker = TeamStatusTracker()
#
#     # Page configuration
#     st.set_page_config(page_title="Team Status Tracker", page_icon="📊", layout="wide")
#
#     # Custom CSS
#     st.markdown("""
#     <style>
#     .main-container {
#         max-width: 1000px;
#         margin: auto;
#         padding: 20px;
#         background-color: #f0f2f6;
#         border-radius: 10px;
#         box-shadow: 0 4px 6px rgba(0,0,0,0.1);
#     }
#     .stButton>button {
#         background-color: #4CAF50;
#         color: white;
#         width: 100%;
#         padding: 10px;
#         border-radius: 5px;
#     }
#     .stTextInput>div>div>input {
#         padding: 10px;
#         border-radius: 5px;
#     }
#     </style>
#     """, unsafe_allow_html=True)
#
#     # Check login status
#     if 'logged_in' not in st.session_state:
#         st.session_state.logged_in = False
#
#     # Sidebar for authentication
#     st.sidebar.title("🔐 Team Status Tracker")
#
#     # Authentication flow
#     if not st.session_state.logged_in:
#         # Choose authentication method
#         auth_method = st.sidebar.radio("Choose Action", ["Login", "Register"])
#
#         if auth_method == "Login":
#             # Login Form
#             with st.sidebar.form("login_form"):
#                 st.header("Login")
#                 login_username = st.text_input("Username")
#                 login_password = st.text_input("Password", type="password")
#                 login_submitted = st.form_submit_button("Login")
#
#                 if login_submitted:
#                     if st.session_state.tracker.validate_login(login_username, login_password):
#                         st.session_state.logged_in = True
#                         st.session_state.username = login_username
#                         st.experimental_rerun()
#                     else:
#                         st.error("Invalid username or password")
#
#         else:
#             # Registration Form
#             with st.sidebar.form("registration_form"):
#                 st.header("Register New Account")
#                 register_username = st.text_input("Choose Username")
#                 register_password = st.text_input("Choose Password", type="password")
#                 register_confirm = st.text_input("Confirm Password", type="password")
#                 register_submitted = st.form_submit_button("Create Account")
#
#                 if register_submitted:
#                     # Validate registration
#                     if register_password != register_confirm:
#                         st.error("Passwords do not match!")
#                     elif len(register_password) < 6:
#                         st.error("Password must be at least 6 characters long")
#                     else:
#                         # Attempt registration
#                         result = st.session_state.tracker.register_user(
#                             register_username,
#                             register_password
#                         )
#                         if result:
#                             st.success("Registration successful! Please login.")
#     else:
#         # Logout functionality
#         st.sidebar.write(f"Logged in as: {st.session_state.username}")
#         if st.sidebar.button("Logout"):
#             st.session_state.logged_in = False
#             st.session_state.username = None
#             st.experimental_rerun()
#
#     # Main content (similar to previous implementation)
#     if st.session_state.logged_in:
#         st.title(f"📊 Daily Status Tracker - {st.session_state.username}")
#
#         # Rest of the application logic remains the same as in previous version
#         # ... (you can keep the existing tabs and status logging code)
#     else:
#         # Welcome screen
#         st.title("🚀 Team Status Tracker")
#         st.write("Please login or register to continue")
#
#
# if __name__ == "__main__":
#     main()


import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime
import os
import hashlib
import sys


class TeamStatusTracker:
    def __init__(self, file_path='team_status_tracker.xlsx'):
        self.file_path = file_path
        self.ensure_file_exists()
        self.users = self.load_users()

    def ensure_file_exists(self):
        """Create the Excel file if it doesn't exist"""
        if not os.path.exists(self.file_path):
            wb = openpyxl.Workbook()
            wb.save(self.file_path)

        # Ensure users sheet exists
        try:
            book = openpyxl.load_workbook(self.file_path)
            if 'Users' not in book.sheetnames:
                users_sheet = book.create_sheet('Users')
                users_sheet['A1'] = 'Username'
                users_sheet['B1'] = 'Password'
                book.save(self.file_path)
        except Exception as e:
            st.error(f"Error creating users sheet: {e}")

    def hash_password(self, password):
        """Hash password for secure storage"""
        return hashlib.sha256(password.encode()).hexdigest()

    def load_users(self):
        """Load users from the Excel file"""
        try:
            book = openpyxl.load_workbook(self.file_path)
            users_sheet = book['Users']

            users = {}
            for row in users_sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    users[row[0]] = row[1]
            return users
        except Exception as e:
            st.error(f"Error loading users: {e}")
            return {}

    def register_user(self, username, password):
        """Register a new user"""
        try:
            # Check for empty username or password
            if not username or not password:
                st.warning("Username and password cannot be empty!")
                return False

            # Check if username already exists
            if username in self.users:
                st.warning("Username already exists!")
                return False

            # Open workbook
            book = openpyxl.load_workbook(self.file_path)
            users_sheet = book['Users']

            # Find next empty row
            next_row = users_sheet.max_row + 1

            # Add new user
            users_sheet.cell(row=next_row, column=1, value=username)
            users_sheet.cell(row=next_row, column=2, value=self.hash_password(password))

            # Save and update
            book.save(self.file_path)
            self.users[username] = self.hash_password(password)

            st.success("User registered successfully!")
            return True
        except Exception as e:
            st.error(f"Error registering user: {e}")
            return False

    def validate_login(self, username, password):
        """Validate user login"""
        if username in self.users:
            return self.users[username] == self.hash_password(password)
        return False


def main():
    # Set up session state
    if 'logged_in' not in st.session_state:
        st.session_state.logged_in = False

    # Initialize tracker
    if 'tracker' not in st.session_state:
        st.session_state.tracker = TeamStatusTracker()

    # Page configuration
    st.set_page_config(page_title="Team Status Tracker", page_icon="📊", layout="wide")

    # Custom CSS
    st.markdown("""
    <style>
    .main-container {
        max-width: 1000px;
        margin: auto;
        padding: 20px;
        background-color: #f0f2f6;
        border-radius: 10px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .stButton>button {
        background-color: #4CAF50;
        color: white;
        width: 100%;
        padding: 10px;
        border-radius: 5px;
    }
    .stTextInput>div>div>input {
        padding: 10px;
        border-radius: 5px;
    }
    </style>
    """, unsafe_allow_html=True)

    # Sidebar for authentication
    st.sidebar.title("🔐 Team Status Tracker")

    # Authentication flow
    if not st.session_state.logged_in:
        # Choose authentication method
        auth_method = st.sidebar.radio("Choose Action", ["Login", "Register"])

        if auth_method == "Login":
            # Login Form
            with st.sidebar.form("login_form", clear_on_submit=True):
                st.header("Login")
                login_username = st.text_input("Username", key="login_username")
                login_password = st.text_input("Password", type="password", key="login_password")
                login_submitted = st.form_submit_button("Login")

                if login_submitted:
                    if st.session_state.tracker.validate_login(login_username, login_password):
                        st.session_state.logged_in = True
                        st.session_state.username = login_username
                        st.rerun()
                    else:
                        st.error("Invalid username or password")

        else:
            # Registration Form
            with st.sidebar.form("registration_form", clear_on_submit=True):
                st.header("Register New Account")
                register_username = st.text_input("Choose Username", key="reg_username")
                register_password = st.text_input("Choose Password", type="password", key="reg_password")
                register_confirm = st.text_input("Confirm Password", type="password", key="reg_confirm")
                register_submitted = st.form_submit_button("Create Account")

                if register_submitted:
                    # Validate registration
                    if register_password != register_confirm:
                        st.error("Passwords do not match!")
                    elif len(register_password) < 6:
                        st.error("Password must be at least 6 characters long")
                    else:
                        # Attempt registration
                        result = st.session_state.tracker.register_user(
                            register_username,
                            register_password
                        )
                        if result:
                            st.success("Registration successful! Please login.")
    else:
        # Logout functionality
        st.sidebar.write(f"Logged in as: {st.session_state.username}")
        if st.sidebar.button("Logout"):
            st.session_state.logged_in = False
            st.session_state.username = None
            st.rerun()

    # Main content
    if st.session_state.logged_in:
        st.title(f"📊 Daily Status Tracker - {st.session_state.username}")

        # Tabs for different functionalities
        tab1, tab2 = st.tabs(["Create Team Member Sheet", "Log Daily Status"])

        with tab1:
            st.header("Create New Team Member Sheet")
            new_member_name = st.text_input("Enter Team Member Name", key="new_member_input")
            if st.button("Create Sheet"):
                if new_member_name:
                    st.session_state.tracker.create_team_member_sheet(
                        new_member_name,
                        st.session_state.username
                    )
                else:
                    st.warning("Please enter a team member name")

        with tab2:
            st.header("Log Daily Status")

            # Load existing sheets for current user
            try:
                book = openpyxl.load_workbook(st.session_state.tracker.file_path)
                user_sheets = [
                    sheet.title for sheet in book.worksheets
                    if sheet.title != 'Users' and
                       sheet.cell(row=1, column=2).value == st.session_state.username
                ]
            except:
                user_sheets = []

            # Member selection
            member_name = st.selectbox("Select Team Member", user_sheets)

            # Status input fields
            tasks_completed = st.text_area("Tasks Completed Today")
            tasks_pending = st.text_area("Tasks Pending")
            blockers = st.text_area("Any Blockers?")
            hours_worked = st.number_input("Hours Worked Today", min_value=0.0, max_value=24.0, step=0.5)

            # Log status button
            if st.button("Log Status"):
                if member_name:
                    st.session_state.tracker.log_daily_status(
                        member_name,
                        st.session_state.username,
                        tasks_completed,
                        tasks_pending,
                        blockers,
                        hours_worked
                    )
                else:
                    st.warning("Please select a team member first")
    else:
        # Welcome screen
        st.title("🚀 Team Status Tracker")
        st.write("Please login or register to continue")


# Add methods to TeamStatusTracker for sheet creation and status logging
def create_team_member_sheet(self, member_name, current_user):
    """Create a new sheet for a team member"""
    try:
        book = openpyxl.load_workbook(self.file_path)

        # Check if sheet already exists
        if member_name in book.sheetnames:
            st.warning(f"Sheet for {member_name} already exists!")
            return False

        # Create new sheet
        sheet = book.create_sheet(title=member_name)

        # Add owner information
        sheet.cell(row=1, column=1, value="Owner")
        sheet.cell(row=1, column=2, value=current_user)

        # Add headers
        headers = ['Date', 'Tasks Completed', 'Tasks Pending', 'Blockers', 'Hours Worked']
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=2, column=col, value=header)

        book.save(self.file_path)
        st.success(f"Sheet created for {member_name}")
        return True
    except Exception as e:
        st.error(f"Error creating sheet: {e}")
        return False


def log_daily_status(self, member_name, current_user, tasks_completed, tasks_pending, blockers, hours_worked):
    """Log daily status for a team member"""
    try:
        book = openpyxl.load_workbook(self.file_path)

        # Check if sheet exists
        if member_name not in book.sheetnames:
            st.error(f"Sheet for {member_name} does not exist. Create it first!")
            return False

        # Get the sheet
        sheet = book[member_name]

        # Verify ownership
        sheet_owner = sheet.cell(row=1, column=2).value
        if sheet_owner != current_user:
            st.error("You do not have permission to log status for this member!")
            return False

        # Find the next row
        next_row = sheet.max_row + 1

        # Write data
        sheet.cell(row=next_row, column=1, value=datetime.now().strftime('%Y-%m-%d'))
        sheet.cell(row=next_row, column=2, value=tasks_completed)
        sheet.cell(row=next_row, column=3, value=tasks_pending)
        sheet.cell(row=next_row, column=4, value=blockers)
        sheet.cell(row=next_row, column=5, value=hours_worked)

        book.save(self.file_path)
        st.success("Daily status logged successfully!")
        return True
    except Exception as e:
        st.error(f"Error logging status: {e}")
        return False


# Add the methods to the class
TeamStatusTracker.create_team_member_sheet = create_team_member_sheet
TeamStatusTracker.log_daily_status = log_daily_status

if __name__ == "__main__":
    main()